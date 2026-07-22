#!/usr/bin/env python3
"""
Convert /login/rates.xlsx → /rates.json

Run from project root:
    python3 scripts/convert-rates.py

The rates.xlsx is a wide pivot table. Each column is one bank+package+qualifying-amount
combination. This script normalises it into a flat JSON array of packages, each with
parsed numeric year-by-year rates, lock-in years, and lender info.

Reference rates (1M SORA, 3M SORA, FHR6) are read from rows 27-28 / inferred from formulas,
and used to compute floating package rates that don't have an explicit "= X.XX%" suffix.
"""
import json
import os
import re
import sys
import glob
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
# rates.xlsx source, newest-first: Rates/ folder, then repo root, then legacy login/.
_XLSX_RATESDIR = ROOT / "Rates" / "rates.xlsx"
_XLSX_ROOT     = ROOT / "rates.xlsx"
_XLSX_LEGACY   = ROOT / "login" / "rates.xlsx"


def _sheet_date(path):
    """Sort key for an advisor sheet, from a filename like
    '13 July 2026 (For Advisors Only) - PTE COMPLETED.xlsx'.

    Sheets arrive dated, so the filename is the freshness signal. The "As at"
    cell inside them is unreliable — it is often left on a previous month's
    date. Falls back to mtime when the name has no parseable date.
    """
    p = Path(path)
    m = re.search(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})', p.name)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)[:3]} {m.group(3)}", "%d %b %Y")
        except ValueError:
            pass
    return datetime.fromtimestamp(os.path.getmtime(p))


def _parse_as_of(text):
    """'13 July 2026' / 'As at 11 May 2026' / 'Apr 2026' -> datetime, or None."""
    if not text:
        return None
    s = re.sub(r'^\s*As\s+(?:of|at)\s+', '', str(text), flags=re.IGNORECASE).strip()
    for fmt in ("%d %B %Y", "%d %b %Y", "%B %Y", "%b %Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def as_of_from_filename(path):
    """'13 July 2026 (...) - PTE COMPLETED.xlsx' -> '13 July 2026'.

    Returns None when the name carries no date (e.g. the fixed rates.xlsx), so
    the caller can fall back to the in-sheet cell.
    """
    m = re.search(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})', Path(path).name)
    if not m:
        return None
    try:
        datetime.strptime(f"{m.group(1)} {m.group(2)[:3]} {m.group(3)}", "%d %b %Y")
    except ValueError:
        return None
    return f"{int(m.group(1))} {m.group(2)} {m.group(3)}"


def dated_sheets(keyword):
    """Advisor sheets under Rates/ whose filename contains `keyword`, newest first.

    Recursive so a sheet dropped into a subfolder is still found, and anything
    filed under archive/ is skipped. Matching is case-insensitive on the name so
    behaviour is identical on case-sensitive filesystems (CI) and macOS.
    """
    hits = [Path(h) for h in glob.glob(str(ROOT / "Rates" / "**" / "*.xlsx"), recursive=True)]
    hits = [h for h in hits
            if "archive" not in h.parts
            and not h.name.startswith("~$")          # Excel lock files
            and keyword.upper() in h.name.upper()]
    hits.sort(key=_sheet_date, reverse=True)
    return hits


def _find_completed():
    """Newest dated COMPLETED sheet, else the fixed rates.xlsx fallbacks.

    Keeps CI working: the dated sheets are gitignored, so in the daily bot's
    checkout there are none and this falls through to the decrypted rates.xlsx.
    """
    hits = dated_sheets("COMPLETED")
    if hits:
        return hits[0]
    return next((p for p in (_XLSX_RATESDIR, _XLSX_ROOT, _XLSX_LEGACY) if p.exists()), _XLSX_LEGACY)


XLSX = _find_completed()
OUT  = ROOT / "rates.json"
HISTORY = ROOT / "rates-history.json"   # append-only audit log of rate changes

# Reference rate fallbacks (used if not detected in the sheet)
DEFAULT_REFS = {"sora1m": 1.076, "sora3m": 1.122, "fhr6": 0.80}

# CLI flag --use-json-refs: prefer refRates from existing rates.json over xlsx.
# Used by the daily MAS cron so live SORA values flow into per-package year1Rate.
USE_JSON_REFS = "--use-json-refs" in sys.argv or os.environ.get("USE_JSON_REFS") == "1"

# CLI flag --force: write even when the source workbook is older than the
# already-published rates.json (see the staleness guard in main()).
FORCE = "--force" in sys.argv

# ---------- helpers ----------

def forward_fill(values):
    """Take a list, propagate the last non-None value forward into None slots."""
    out = []
    last = None
    for v in values:
        if v is not None and str(v).strip():
            last = v
        out.append(last)
    return out

def parse_qualifying_min(text):
    """`>$1.5M`, `> $500K`, `> $1M + $30K Privilege onboard + CC` → 1500000 / 500000 / 1000000."""
    if not text: return None
    m = re.search(r'\$\s*([\d.]+)\s*(M|K|mil|k)', str(text), re.IGNORECASE)
    if not m: return None
    val = float(m.group(1))
    unit = m.group(2).upper()
    if unit in ('M', 'MIL'): return int(val * 1_000_000)
    if unit == 'K':          return int(val * 1_000)
    return None

def parse_lockin(text):
    """`2yrs` / `2 yrs` / `1yr` / `3yrs` → 2 / 2 / 1 / 3."""
    if not text: return None
    m = re.search(r'(\d+)\s*yr', str(text), re.IGNORECASE)
    return int(m.group(1)) if m else None

def parse_rate(text, refs):
    """
    Parse a year-by-year rate cell into a numeric % value.

    Handles patterns:
      'FHR6 (0.80%) + 0.55% = 1.35%'        → 1.35  (use after =)
      'FHR 6 (1.40%) + 0.70%'               → 2.10  (compute from base + spread inline)
      '1.45% Fixed' / '1.45 Fixed'          → 1.45
      '1M SORA + 0.00% = 1.076%'            → 1.076
      '1M SORA + 0.60%'                     → refs.sora1m + 0.60
      '3M SORA + 0.30% = 1.422%'            → 1.422
      '1M/3M SORA + 0.55%'                  → refs.sora3m + 0.55  (use the slower of the two)
      '1.50% Fixed' or any 'X.XX% Fixed'    → 1.50
    """
    if text is None: return None
    s = str(text).strip()
    if not s: return None

    # 1) Has '= X.XX%' suffix → use that
    m = re.search(r'=\s*([\d.]+)\s*%?', s)
    if m:
        try: return round(float(m.group(1)), 4)
        except: pass

    # 2) Pattern 'FHR<n> (X.XX%) + Y.YY%'  — compute base + spread
    m = re.search(r'FHR\s*\d?\s*\(\s*([\d.]+)\s*%\s*\)\s*\+\s*([\d.]+)\s*%', s, re.IGNORECASE)
    if m:
        try: return round(float(m.group(1)) + float(m.group(2)), 4)
        except: pass

    # 3) 'X.XX% Fixed' or 'X.XX Fixed'
    m = re.search(r'([\d.]+)\s*%?\s*Fixed', s, re.IGNORECASE)
    if m:
        try: return round(float(m.group(1)), 4)
        except: pass

    # 4) '1M SORA + X.XX%' / '3M SORA + X.XX%' / '1M/3M SORA + X.XX%'
    m_sora = re.search(r'(1M|3M|1M\s*/\s*3M)\s*SORA\s*\+\s*([\d.]+)\s*%', s, re.IGNORECASE)
    if m_sora:
        ref_label = m_sora.group(1).upper().replace(' ', '')
        try:
            spread = float(m_sora.group(2))
            if ref_label == '1M':           ref = refs['sora1m']
            elif ref_label == '3M':         ref = refs['sora3m']
            elif ref_label == '1M/3M':      ref = max(refs['sora1m'], refs['sora3m'])
            else:                            ref = refs['sora3m']
            return round(ref + spread, 4)
        except: pass

    # 5) Bare 'FHR' reference without parens → use ref + spread
    m = re.search(r'FHR\s*\d?\s*\+\s*([\d.]+)\s*%', s, re.IGNORECASE)
    if m:
        try: return round(refs['fhr6'] + float(m.group(1)), 4)
        except: pass

    # 6) Fallback: first percentage anywhere
    m = re.search(r'([\d.]+)\s*%', s)
    if m:
        try: return round(float(m.group(1)), 4)
        except: pass

    return None

def parse_ref_rate(label_text):
    """`1M COMPOUNDED SORA: 1.076%` → 1.076"""
    if not label_text: return None
    m = re.search(r':\s*([\d.]+)\s*%', str(label_text))
    return float(m.group(1)) if m else None

# ---------- main ----------

def main():
    # Log the chosen source: silently parsing the wrong (stale) sheet is the
    # failure mode that matters here, so make the pick visible every run.
    _all = dated_sheets("COMPLETED")
    if len(_all) > 1:
        print(f"[convert-rates] {len(_all)} completed sheets found; using newest")
    print(f"[convert-rates] source: {XLSX.relative_to(ROOT) if XLSX.is_relative_to(ROOT) else XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["All Rates"]

    # 1) Reference rates
    refs = dict(DEFAULT_REFS)
    r27_label = ws.cell(27, 1).value
    r28_label = ws.cell(28, 1).value
    v = parse_ref_rate(r27_label)
    if v is not None: refs['sora1m'] = v
    v = parse_ref_rate(r28_label)
    if v is not None: refs['sora3m'] = v

    # 1a) If --use-json-refs flag set, override with values from existing rates.json.
    # Lets the daily MAS cron flow live SORA into per-package year1Rate.
    if USE_JSON_REFS and OUT.exists():
        try:
            existing = json.loads(OUT.read_text())
            json_refs = existing.get("refRates") or {}
            for k in ("sora1m", "sora3m", "sora6m", "fhr6"):
                if k in json_refs and json_refs[k] is not None:
                    refs[k] = float(json_refs[k])
            print(f"[convert-rates] --use-json-refs: refs from rates.json = {refs}")
        except Exception as e:
            print(f"[convert-rates] WARN: --use-json-refs requested but failed: {e}")

    # As-of date — best-effort: row 26 col 1 might say "As of 27 Feb 2026" or "As at 11 May 2026"
    # Filename date wins over the in-sheet "As at" cell. The advisor sheets ship
    # with that cell left on an earlier month (the 13 Jul 2026 sheet still said
    # "As at 11 May 2026"), and publishing July pricing under a May date is worse
    # than useless on a rates page. Fall back to the cell only for the fixed
    # rates.xlsx, which carries no date in its name.
    as_of = as_of_from_filename(XLSX)
    if not as_of:
        for r in range(24, 30):
            cell = ws.cell(r, 1).value
            if cell:
                m = re.match(r'\s*As\s+(?:of|at)\s+(.+)', str(cell), re.IGNORECASE)
                if m:
                    as_of = m.group(1).strip()
                    break
    if not as_of:
        as_of = "Apr 2026"

    # 2) Forward-fill the header rows (category, subcat, lender, lock-in)
    last_col = ws.max_column
    cats     = forward_fill([ws.cell(1, c).value for c in range(2, last_col + 1)])
    subcats  = forward_fill([ws.cell(2, c).value for c in range(2, last_col + 1)])
    lenders  = [ws.cell(4, c).value for c in range(2, last_col + 1)]  # do NOT forward-fill lender — multi-lender groups
    qualifs  = [ws.cell(3, c).value for c in range(2, last_col + 1)]
    lockins  = forward_fill([ws.cell(12, c).value for c in range(2, last_col + 1)])

    packages = []
    for idx, c in enumerate(range(2, last_col + 1)):
        lender = lenders[idx]
        if not lender:
            continue  # skip empty / spacer columns
        cat    = cats[idx] or ''
        sub    = subcats[idx] or ''
        qual   = qualifs[idx] or ''
        lockin = lockins[idx] or ''
        # parse year-by-year
        yr1 = ws.cell(5, c).value
        yr2 = ws.cell(6, c).value
        yr3 = ws.cell(7, c).value
        yr4 = ws.cell(8, c).value
        yr5 = ws.cell(9, c).value
        thr = ws.cell(10, c).value
        rate1 = parse_rate(yr1, refs)
        # Skip rows where year-1 has no rate (likely a malformed column)
        if rate1 is None:
            continue
        # Determine product family for /free-report/ matching
        family = None
        sub_l = str(sub).lower()
        cat_l = str(cat).lower()
        if 'fhr' in sub_l:               family = 'fhr'
        elif '1 mth sora' in sub_l or '1m sora' in sub_l: family = 'sora1m'
        elif '3 mths sora' in sub_l or '3m sora' in sub_l: family = 'sora3m'
        elif 'combo' in sub_l or 'combo' in cat_l:        family = 'combo'
        elif '1yr fixed' in sub_l or '1 yr fixed' in sub_l or '1year fixed' in sub_l: family = 'fixed1y'
        elif '2yrs fixed' in sub_l or '2 yrs fixed' in sub_l: family = 'fixed2y'
        elif '3yrs fixed' in sub_l or '3 yrs fixed' in sub_l: family = 'fixed3y'

        pkg = {
            "id": idx + 1,
            "lender": str(lender).strip(),
            "category": str(cat).replace(' RATES', '').strip().title(),  # 'Variable', 'Combo', 'Fixed'
            "subCategory": str(sub).strip(),
            "family": family,
            "qualifying": str(qual).strip(),
            "qualifyingMin": parse_qualifying_min(qual),
            "lockInYears": parse_lockin(lockin),
            "year1Rate": rate1,
            "year2Rate": parse_rate(yr2, refs),
            "year3Rate": parse_rate(yr3, refs),
            "year4Rate": parse_rate(yr4, refs),
            "year5Rate": parse_rate(yr5, refs),
            "thereafterRate": parse_rate(thr, refs),
            "year1Raw": (str(yr1).strip() if yr1 else None),
            "year2Raw": (str(yr2).strip() if yr2 else None),
            "year3Raw": (str(yr3).strip() if yr3 else None),
            "year4Raw": (str(yr4).strip() if yr4 else None),
            "year5Raw": (str(yr5).strip() if yr5 else None),
            "thereafterRaw": (str(thr).strip() if thr else None),
        }
        packages.append(pkg)

    out = {
        "asOf": as_of,
        "refRates": refs,
        "packages": packages,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "xlsxFingerprint": _hash_xlsx(XLSX),
    }

    # ---- Diff against previous rates.json (for history) ----
    prev = None
    if OUT.exists():
        try:
            prev = json.loads(OUT.read_text())
        except Exception as e:
            print(f"  (could not read previous rates.json: {e})")

    # Preserve MAS-API-sourced asOfSora written by fetch-mas-rates.mjs.
    # convert-rates.py overwrites rates.json from xlsx; without this guard,
    # the daily MAS asOfSora gets clobbered every run and the front end
    # falls back to the stale "As of …" string typed into the xlsx.
    # SORA numeric values are already preserved via --use-json-refs flag.
    if prev:
        prev_as_of_sora = prev.get("asOfSora")
        if prev_as_of_sora:
            out["asOfSora"] = prev_as_of_sora
        prev_inner = (prev.get("refRates") or {}).get("asOfSora")
        if prev_inner:
            out["refRates"]["asOfSora"] = prev_inner

    # ---- Staleness guard ----
    # Refuse to overwrite newer published rates with an older workbook. The CI
    # job decrypts rates.xlsx.enc, so whenever that blob lags behind a sheet that
    # was published from a workstation, the nightly run would silently roll the
    # site back to the older pricing. Skip instead of failing: the SORA refresh
    # in the surrounding job is still valid and should complete.
    if prev and not FORCE:
        prev_dt = _parse_as_of(prev.get("asOf"))
        cur_dt  = _parse_as_of(as_of)
        if prev_dt and cur_dt and cur_dt < prev_dt:
            print(f"::warning::[convert-rates] SKIPPED write — source sheet is older than published rates.")
            print(f"  source   : {XLSX.name} (asOf {as_of})")
            print(f"  published: rates.json (asOf {prev.get('asOf')})")
            print(f"  rates.json left untouched. If this is CI, rates.xlsx.enc needs re-encrypting")
            print(f"  from the newer sheet. Override with --force.")
            return

    diffs = _diff_packages(prev["packages"] if prev else [], packages) if prev else []
    refDiffs = _diff_refs(prev.get("refRates") if prev else {}, refs)

    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"✓ Wrote {len(packages)} packages → {OUT}")

    # ---- Append history entry ----
    entry = {
        "generatedAt": out["generatedAt"],
        "asOf": as_of,
        "xlsxFingerprint": out["xlsxFingerprint"],
        "packageCount": len(packages),
        "refRates": refs,
        "refRateChanges": refDiffs,
        "packageChanges": diffs,           # only entries where year1Rate / lockIn / qualifyingMin differs
        "summary": _summarise_diff(diffs, refDiffs, prev is not None),
    }
    history = []
    if HISTORY.exists():
        try:
            history = json.loads(HISTORY.read_text())
        except Exception:
            history = []
    history.append(entry)
    # Keep only last 200 entries
    history = history[-200:]
    HISTORY.write_text(json.dumps(history, indent=2, ensure_ascii=False))
    print(f"✓ Appended history entry → {HISTORY} ({len(history)} total)")

    # ---- Sanity summary ----
    by_family = {}
    by_lender = {}
    for p in packages:
        by_family[p['family']] = by_family.get(p['family'], 0) + 1
        by_lender[p['lender']] = by_lender.get(p['lender'], 0) + 1
    print(f"  Families: {by_family}")
    print(f"  Lenders:  {by_lender}")
    print(f"  Diff: {entry['summary']}")


def _hash_xlsx(path):
    """SHA-1 fingerprint of the xlsx so we can tell whether the master file changed."""
    if not path.exists():
        return None
    h = hashlib.sha1()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()[:12]


def _pkg_key(p):
    """Stable key per (lender, family, qualifyingMin) — survives row reordering in xlsx."""
    return f"{p.get('lender','?')}|{p.get('family','?')}|{p.get('qualifyingMin') or 0}"


def _diff_packages(old, new):
    """Return a list of {key, what, before, after} for material changes."""
    om = {_pkg_key(p): p for p in old}
    nm = {_pkg_key(p): p for p in new}
    changes = []
    for key, np in nm.items():
        op = om.get(key)
        if op is None:
            changes.append({
                "key": key, "lender": np['lender'], "package": np.get('subCategory'),
                "qualifying": np.get('qualifying'), "what": "added", "after": np.get('year1Rate')
            })
            continue
        if op.get('year1Rate') != np.get('year1Rate'):
            delta = round((np.get('year1Rate') or 0) - (op.get('year1Rate') or 0), 4)
            changes.append({
                "key": key, "lender": np['lender'], "package": np.get('subCategory'),
                "qualifying": np.get('qualifying'),
                "what": "year1Rate",
                "before": op.get('year1Rate'),
                "after": np.get('year1Rate'),
                "deltaPct": delta,
            })
        if op.get('lockInYears') != np.get('lockInYears'):
            changes.append({
                "key": key, "lender": np['lender'],
                "what": "lockInYears",
                "before": op.get('lockInYears'),
                "after": np.get('lockInYears'),
            })
        if op.get('qualifyingMin') != np.get('qualifyingMin'):
            changes.append({
                "key": key, "lender": np['lender'],
                "what": "qualifyingMin",
                "before": op.get('qualifyingMin'),
                "after": np.get('qualifyingMin'),
            })
    for key, op in om.items():
        if key not in nm:
            changes.append({
                "key": key, "lender": op['lender'], "package": op.get('subCategory'),
                "qualifying": op.get('qualifying'), "what": "removed", "before": op.get('year1Rate')
            })
    return changes


def _diff_refs(old, new):
    out = {}
    for k in ('sora1m', 'sora3m', 'fhr6'):
        ov = old.get(k) if old else None
        nv = new.get(k)
        if ov != nv:
            out[k] = {"before": ov, "after": nv,
                      "deltaPct": round((nv or 0) - (ov or 0), 4) if ov is not None else None}
    return out


def _summarise_diff(pkgChanges, refChanges, hadPrevious):
    if not hadPrevious:
        return "Initial rates.json — no prior version to diff against."
    if not pkgChanges and not refChanges:
        return "No material changes."
    parts = []
    if refChanges:
        parts.append(f"ref rates changed: {', '.join(refChanges.keys())}")
    by_kind = {}
    for c in pkgChanges:
        by_kind[c['what']] = by_kind.get(c['what'], 0) + 1
    parts.append(", ".join(f"{n} {k}" for k, n in by_kind.items()))
    return " · ".join(parts)


if __name__ == "__main__":
    main()

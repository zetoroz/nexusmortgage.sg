#!/usr/bin/env python3
"""
Convert the BUC (Building-Under-Construction) private-property rate sheet
-> /rates-buc.json, the public BUC counterpart of /rates.json.

The BUC sheet has a +1 row offset vs the Completed "All Rates" sheet (an extra
"Special Highlights" row): lenders on row 5, year rates on rows 6-11, lock-in on
row 15. Category is derived from the sub-type (the sheet's row-1 title is just
"BUC PRIVATE PROPERTY HOME LOANS").

Reuses convert-rates.py's rate parsers (no duplication, no divergence — its
main() is __main__-guarded so importing it runs nothing). SORA refs are taken
from rates.json so floating BUC rates track the same live SORA.

Run from repo root:  python3 scripts/convert-rates-buc.py
Source: Rates/<...BUC...>.xlsx  (private, gitignored — same as the Completed sheet).
"""
import json, re, os, glob, importlib.util
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# Reuse the canonical parsers from convert-rates.py.
_spec = importlib.util.spec_from_file_location("convert_rates", str(ROOT / "scripts" / "convert-rates.py"))
cr = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(cr)

OUT = ROOT / "rates-buc.json"
RATES_JSON = ROOT / "rates.json"


def find_buc():
    """Newest BUC sheet, searched recursively.

    Was: non-recursive glob + hits[0]. That silently missed sheets dropped into
    a subfolder, and with more than one BUC file present it picked whichever the
    filesystem happened to list first — which could be a stale sheet. On a rates
    page that means publishing outdated pricing, so pick the newest explicitly.
    Selection is shared with convert-rates.py via cr.dated_sheets.
    """
    hits = cr.dated_sheets("BUC") + [Path(h) for h in glob.glob(str(ROOT / "*BUC*.xlsx"))]
    if not hits:
        return None
    if len(hits) > 1:
        print(f"[buc] {len(hits)} BUC sheets found; using newest: {os.path.basename(str(hits[0]))}")
    return str(hits[0])


def category_from_sub(sub):
    s = str(sub).lower()
    if "fixed" in s:  return "Fixed"
    if "combo" in s:  return "Combo"
    return "Variable"   # SORA / FHR pegged


def raw(v):
    return str(v).strip() if v not in (None, "") else ""


def main():
    src = find_buc()
    if not src:
        print("[buc] No BUC xlsx found in Rates/ — nothing to do.")
        return

    # Live SORA / FHR refs from rates.json (else defaults)
    refs = dict(cr.DEFAULT_REFS)
    if RATES_JSON.exists():
        try:
            j = json.loads(RATES_JSON.read_text()).get("refRates") or {}
            for k in ("sora1m", "sora3m", "sora6m", "fhr6"):
                if j.get(k) is not None:
                    refs[k] = float(j[k])
        except Exception as e:
            print("[buc] WARN refs from rates.json:", e)

    wb = load_workbook(src, data_only=True)
    ws = wb["BUC"]
    last = ws.max_column

    # Filename date wins over the in-sheet "As at" cell, which the advisor sheets
    # routinely leave on an earlier month (the 1 Jul 2026 BUC sheet still said
    # "As at 11 May 2026"). Cell is the fallback for sheets with no dated name.
    as_of = cr.as_of_from_filename(src)
    if not as_of:
        for r in range(1, 30):
            c = ws.cell(r, 1).value
            if c:
                m = re.match(r"\s*As\s+(?:of|at)\s+(.+)", str(c), re.IGNORECASE)
                if m:
                    as_of = m.group(1).strip()
                    break

    subcats = cr.forward_fill([ws.cell(2, c).value for c in range(2, last + 1)])
    quals   = [ws.cell(3, c).value for c in range(2, last + 1)]
    lenders = [ws.cell(5, c).value for c in range(2, last + 1)]            # BUC: lenders row 5
    lockins = cr.forward_fill([ws.cell(15, c).value for c in range(2, last + 1)])  # BUC: lock-in row 15

    packages = []
    for idx, c in enumerate(range(2, last + 1)):
        lender = lenders[idx]
        if not lender:
            continue
        sub = subcats[idx] or ""
        qual = quals[idx] or ""
        lockin = lockins[idx] or ""
        yr = [ws.cell(r, c).value for r in (6, 7, 8, 9, 10, 11)]   # BUC: 1st..5th + thereafter on rows 6-11
        rate1 = cr.parse_rate(yr[0], refs)
        if rate1 is None:
            continue

        sl = str(sub).lower()
        family = None
        if "fhr" in sl:                                   family = "fhr"
        elif "1 mth sora" in sl or "1m sora" in sl:       family = "sora1m"
        elif "3 mths sora" in sl or "3m sora" in sl:      family = "sora3m"
        elif "combo" in sl:                               family = "combo"
        elif "1yr fixed" in sl or "1 yr fixed" in sl:     family = "fixed1y"
        elif "2yrs fixed" in sl or "2 yrs fixed" in sl:   family = "fixed2y"
        elif "3yrs fixed" in sl or "3 yrs fixed" in sl:   family = "fixed3y"

        packages.append({
            "id": idx + 1,
            "lender": str(lender).strip(),
            "category": category_from_sub(sub),
            "subCategory": str(sub).strip(),
            "family": family,
            "qualifying": str(qual).strip(),
            "qualifyingMin": cr.parse_qualifying_min(qual),
            "lockInYears": cr.parse_lockin(lockin),
            "lockInRaw": raw(lockin),
            "year1Rate": rate1,
            "year2Rate": cr.parse_rate(yr[1], refs),
            "year3Rate": cr.parse_rate(yr[2], refs),
            "year4Rate": cr.parse_rate(yr[3], refs),
            "year5Rate": cr.parse_rate(yr[4], refs),
            "thereafterRate": cr.parse_rate(yr[5], refs),
            "year1Raw": raw(yr[0]), "year2Raw": raw(yr[1]), "year3Raw": raw(yr[2]),
            "year4Raw": raw(yr[3]), "year5Raw": raw(yr[4]), "thereafterRaw": raw(yr[5]),
            "propertyStatus": "buc",
        })

    out = {
        "asOf": as_of or "",
        "propertyStatus": "buc",
        "refRates": refs,
        "packages": packages,
    }
    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n")
    print(f"[buc] Wrote {len(packages)} BUC packages -> {OUT}  (asOf {as_of})")


if __name__ == "__main__":
    main()
